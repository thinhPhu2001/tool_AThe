from config import *

import openpyxl
from tqdm import tqdm
import win32com.client as win32
import matplotlib.pyplot as plt
import pandas as pd
from time import sleep
import pyperclip
import os
import datetime
from PIL import Image
from openpyxl import load_workbook
from datetime import datetime, timedelta
import pyautogui
import win32gui
import win32con
import threading
import logging
import time

Image.MAX_IMAGE_PIXELS = 300000000
stop_thread = threading.Event()


class ExcelManager:  # win32
    def __init__(self, file_path):
        """
        Khởi tạo và mở file Excel.

        Args:
            file_path (str): Đường dẫn tới file Excel.
        """
        self.file_path = file_path
        self.excel = None
        self.workbook = None
        # self.open_file()

    def open_file(self):
        """
        Mở file Excel và khởi tạo đối tượng workbook nếu chưa mở.
        """
        if self.is_file_open():
            print(f"File Excel '{self.file_path}' đã được mở trước đó.")
            return True

        print("Đang mở file Excel...")
        try:
            self.excel = win32.Dispatch("Excel.Application")
            self.workbook = self.excel.Workbooks.Open(self.file_path)
            self.excel.DisplayAlerts = False
            print(f"File Excel '{self.file_path}' đã được mở thành công.")
            return True
        except Exception as e:
            print(f"Lỗi khi mở file Excel: {e}")
            self.excel = None
            self.workbook = None
            return False

    def is_file_open(self):
        """
        Kiểm tra xem file Excel đã được mở trong ứng dụng Excel hay chưa.
        """
        try:
            if self.excel is not None:
                for wb in self.excel.Workbooks:
                    if wb.FullName == self.file_path:
                        return True
            return False
        except Exception as e:
            print(f"Lỗi khi kiểm tra file đã mở: {e}")
            return False

    def save_file(self, save_path=None):
        """
        Lưu file Excel. Nếu không chỉ định `save_path`, file sẽ được lưu đè.
        Args:
            save_path (str, optional): Đường dẫn lưu file mới.
        """
        if self.workbook is None:
            raise Exception("Workbook chưa được mở. Vui lòng gọi `open_excel()` trước.")

        if save_path:
            # Tạo thư mục nếu chưa tồn tại
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            self.workbook.SaveAs(save_path)
        else:
            self.workbook.Save()  # Lưu đè file gốc
            if save_path is not None:
                print(f"Đã lưu xuống: {save_path}")
            else:
                print(f"Đã lưu xuống: {self.file_path}")

    def close_all_file(self):
        """
        Đóng file Excel và thoát Excel Application. (đóng tất cả, kể cả các file Excel mở trước đó)
        """
        # Đặt clipboard của Excel về rỗng để tránh xung đột khi thoát Excel
        pyperclip.copy("")

        # Đóng workbook nếu tồn tại
        if self.workbook:
            try:
                # Lưu workbook trước khi đóng (nếu cần)
                self.workbook.Save()
                # Đặt trạng thái "Saved" để Excel không hỏi xác nhận
                self.workbook.Saved = True
                # Đóng workbook
                self.workbook.Close(SaveChanges=0)
                print("File Excel đã được đóng.")
            except Exception as e:
                print(f"Lỗi khi đóng file Excel: {e}")
            finally:
                self.workbook = None  # Xóa tham chiếu đến workbook

        # Thoát ứng dụng Excel
        if self.excel:
            try:
                self.excel.DisplayAlerts = False  # tắt thông báo excel
                self.excel.Application.Quit()  # Thoát ứng dụng Excel
                print("Excel Application đã được thoát.")
            except Exception as e:
                print(f"Lỗi khi thoát Excel: {e}")
            finally:
                self.excel = None  # Xóa tham chiếu đến Excel Application

    def get_sheet(self, sheet_name):
        """
        Lấy đối tượng sheet từ workbook.

        Args:
            sheet_name (str): Tên sheet cần lấy.

        Returns:
            sheet: Đối tượng sheet hoặc None nếu không tìm thấy.
        """
        try:
            for sheet in self.workbook.Sheets:
                if sheet.Name == sheet_name:
                    return sheet
            print(f"Không tìm thấy sheet '{sheet_name}'.")
            return None
        except Exception as e:
            print(f"Lỗi khi lấy sheet: {e}")
            return None

    def find_cell_position(self, sheet_name, search_value):
        """
        Tìm vị trí ô có giá trị cụ thể trong sheet.

        Args:
            sheet_name (str): Tên sheet cần đọc.
            search_value (str): Giá trị cần tìm.

        Returns:
            tuple: (row, column) vị trí ô hoặc None nếu không tìm thấy.
        """
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return None

        try:
            for row in range(1, sheet.UsedRange.Rows.Count + 1):
                for col in range(1, sheet.UsedRange.Columns.Count + 1):
                    cell_value = sheet.Cells(row, col).Value
                    if (
                        cell_value is not None
                        and str(cell_value).strip() == search_value
                    ):
                        print(
                            f"Tìm thấy giá trị '{search_value}' tại: (row: {row}, column: {col})"
                        )
                        return (row, col)
            print(
                f"Không tìm thấy giá trị '{search_value}' trong sheet '{sheet_name}'."
            )
            return None
        except Exception as e:
            print(f"Lỗi khi tìm ô: {e}")
            return None

    def get_messages_dep_recipient(self, sheet_name, recipient, message_header):
        """
        Từ người nhận, lấy được tin nhắn tương ứng.

        Args:
            sheet_name (str): Tên sheet cần đọc.
            recipient (str): Tên người nhận.
            message_header (str): Cột chứa tin nhắn.

        Returns:
            str: Tin nhắn tương ứng hoặc None nếu không tìm thấy.
        """
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return None

        try:
            recipient_position = self.find_cell_position(sheet_name, recipient)
            if recipient_position is None:
                print(f"Không tìm thấy '{recipient}' trong sheet '{sheet_name}'")
                return None

            row_recipient, _ = recipient_position
            message_header_position = self.find_cell_position(
                sheet_name, message_header
            )
            if message_header_position is None:
                print(f"Không tìm thấy '{message_header}' trong sheet '{sheet_name}'.")
                return None

            _, col_message = message_header_position
            message = sheet.Cells(row_recipient, col_message).Value
            if message:
                return str(message).strip()
            else:
                print(f"Không có tin nhắn tại dòng {row_recipient}, cột {col_message}.")
                return None
        except Exception as e:
            print(f"Lỗi khi lấy tin nhắn: {e}")
            return None

    def run_macro(self, macro_name):
        """
        Chạy lệnh macro.

        Args:
            macro_name (str): Tên macro muốn chạy.
        """

        def turn_off_Macro_error():
            sleep(3)
            excel_logo = str(IMAGE_PATH / "excel_logo.png")
            end_img = str(IMAGE_PATH / "end_macro.png")

            while not stop_thread.is_set():
                # Tìm hình ảnh Excel Logo
                try:
                    excel_button = pyautogui.locateOnScreen(excel_logo, confidence=0.8)
                    if excel_button:
                        pyautogui.click(pyautogui.center(excel_button))
                        # Tìm nút End Macro

                        try:
                            sleep(2)
                            end_button = pyautogui.locateOnScreen(
                                end_img, confidence=0.6
                            )
                            if end_button:
                                pyautogui.click(
                                    pyautogui.center(end_button)
                                )  # Click vào nút "End Macro"
                                print("Đã xử lý Macro Error.")
                                break
                        except pyautogui.ImageNotFoundException:
                            print("Không tìm thấy hình ảnh END button")

                except pyautogui.ImageNotFoundException:
                    print("Không tìm thấy hình ảnh Excel logo")

                # Nếu không tìm thấy, thử lại sau 10 giây
                print("Không tìm thấy lỗi. Tiếp tục quét...")
                sleep(3)

        stop_thread.clear()
        x = threading.Thread(target=turn_off_Macro_error)
        x.start()

        try:
            self.excel.Application.Run(macro_name)
            print(f"Macro '{macro_name}' đã được chạy thành công!")
            stop_thread.set()
            x.join()
            return True
        except Exception as e:
            print(f"Lỗi khi chạy macro '{macro_name}': {e}. Dừng tiếng trình!")
            stop_thread.set()
            x.join()
            return False

    def copy_data(self, sheet_name, start_row, start_col, end_col):
        """
        Sao chép dữ liệu từ một vùng trong file Excel.

        Args:
            sheet_name (str): Tên sheet cần sao chép dữ liệu.
            start_row (int): Hàng bắt đầu sao chép.
            start_col (int): Cột bắt đầu sao chép.
            end_col (int): Cột kết thúc sao chép.

        Returns:
            source_range: Vùng dữ liệu được sao chép.
        """
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return None

            start_col_num = excel_column_to_number(start_col)
            end_col_num = excel_column_to_number(end_col)

            # Xác định hàng cuối cùng
            last_row = sheet.Cells(sheet.Rows.Count, start_col).End(-4162).Row  # xlUp
            # Xác định vùng dữ liệu
            source_range = sheet.Range(
                sheet.Cells(start_row, start_col_num),
                sheet.Cells(last_row, end_col_num),
            )
            print(
                f"Sao chép dữ liệu từ sheet '{sheet_name}', vùng ({start_row}, {start_col_num}) đến ({last_row}, {end_col_num})."
            )
            return source_range
        except Exception as e:
            print(f"Lỗi khi sao chép dữ liệu: {e}")
            return None

    def copy_result_data(self, sheet_name, start_row, start_col, end_col):
        """
        chỉ sao chép giá trị không sao chép hàm

        Args:
            sheet_name (str): Tên sheet cần sao chép dữ liệu.
            start_row (int): Hàng bắt đầu sao chép.
            start_col (int): Cột bắt đầu sao chép.
            end_col (int): Cột kết thúc sao chép.

        Returns:
            source_range: Vùng dữ liệu được sao chép.
        """
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return None

            start_col_num = excel_column_to_number(start_col)
            end_col_num = excel_column_to_number(end_col)

            # Xác định hàng cuối cùng
            last_row = sheet.Cells(sheet.Rows.Count, start_col).End(-4162).Row  # xlUp

            if last_row < start_row:
                last_row = start_row

            # Xác định vùng dữ liệu
            source_range = sheet.Range(
                sheet.Cells(start_row, start_col_num),
                sheet.Cells(last_row, end_col_num),
            )

            # Sao chép chỉ giá trị (không sao chép công thức)
            values = source_range.Value

            print(
                f"Sao chép dữ liệu từ sheet '{sheet_name}', vùng ({start_row}, {start_col_num}) đến ({last_row}, {end_col_num})."
            )
            return values
        except Exception as e:
            print(f"Lỗi khi sao chép dữ liệu: {e}")
            return None

    def paste_data(self, source_range, dest_sheet_name, dest_start_row, dest_start_col):
        """
        Dán dữ liệu vào một vùng trong file Excel.

        Args:
            source_range: Vùng dữ liệu được sao chép.
            dest_sheet_name (str): Tên sheet cần dán dữ liệu.
            dest_start_row (int): Hàng bắt đầu dán.
            dest_start_col (int): Cột bắt đầu dán.
        """
        dest_start_col_num = excel_column_to_number(dest_start_col)
        try:
            dest_sheet = self.get_sheet(dest_sheet_name)
            if not dest_sheet:
                return None

            dest_cell = dest_sheet.Cells(dest_start_row, dest_start_col_num)
            source_range.Copy(Destination=dest_cell)
            print(
                f"Dán dữ liệu vào sheet '{dest_sheet_name}', ô bắt đầu ({dest_start_row}, {dest_start_col_num})."
            )
        except Exception as e:
            print(f"Lỗi khi dán dữ liệu: {e}")

    def paste_result_data(
        self, source_values, dest_sheet_name, dest_start_row, dest_start_col
    ):
        """
        Dán dữ liệu (giá trị) vào một vùng trong file Excel.

        Args:
            source_values (list): Danh sách giá trị hai chiều được sao chép từ một vùng dữ liệu.
            dest_sheet_name (str): Tên sheet cần dán dữ liệu.
            dest_start_row (int): Hàng bắt đầu dán.
            dest_start_col (int): Cột bắt đầu dán.
        """
        dest_start_col_num = excel_column_to_number(dest_start_col)
        try:
            dest_sheet = self.get_sheet(dest_sheet_name)
            if not dest_sheet:
                return None

            if source_values is None:
                return None

            # Kiểm tra và chuẩn hóa dữ liệu
            if not isinstance(source_values, list) or not all(
                isinstance(row, list) for row in source_values
            ):
                if isinstance(source_values, (int, float)):
                    source_values = [[source_values]]  # Chuyển thành danh sách 2D 1x1
                else:
                    raise ValueError(
                        "source_values phải là danh sách hoặc một giá trị số"
                    )

            num_rows = len(source_values)
            num_cols = len(source_values[0]) if num_rows > 0 else 0

            dest_range = dest_sheet.Range(
                dest_sheet.Cells(dest_start_row, dest_start_col_num),
                dest_sheet.Cells(
                    dest_start_row + num_rows - 1, dest_start_col_num + num_cols - 1
                ),
            )

            # Dán dữ liệu (giá trị)
            dest_range.Value = source_values
            print(
                f"Dán dữ liệu vào sheet '{dest_sheet_name}', vùng ({dest_start_row}, {dest_start_col_num}) đến ({dest_start_row}, {dest_start_col_num + num_cols - 1})."
            )
        except Exception as e:
            print(f"Lỗi khi dán dữ liệu: {e}")

    def clear_data(self, sheet_name, start_row, start_col, end_col):
        """
        Xóa dữ liệu trong sheet, từ cột bắt đầu đến cột kết thúc, từ dòng bắt đầu đến dòng cuối cùng.

        Args:
            sheet (Worksheet): Sheet cần xóa dữ liệu.
            start_row (int): Dòng bắt đầu xóa.
            start_col (str): Cột bắt đầu xóa (ký tự, ví dụ: 'A').
            end_col (str): Cột kết thúc xóa (ký tự, ví dụ: 'Z').
        """
        sheet = self.get_sheet(sheet_name)
        # Tìm dòng cuối cùng trong sheet
        last_row = (
            sheet.Cells(sheet.Rows.Count, start_col).End(-4162).Row
        )  # xlUp = -4162
        if start_row < last_row:
            start_col_num = excel_column_to_number(start_col)
            end_col_num = excel_column_to_number(end_col)
            # Duyệt qua từng dòng và từng cột trong phạm vi cần xóa
            source_range = sheet.Range(
                sheet.Cells(start_row, start_col_num),
                sheet.Cells(last_row, end_col_num),
            )
            if source_range is not None and source_range.Address != "":
                source_range.ClearContents()
                print(
                    f"Đã xóa dữ liệu từ dòng {start_row} đến {last_row}, từ cột {start_col} đến {end_col}."
                )
            else:
                print("Không có dữ liệu trong phạm vi cần xóa.")
        else:
            print(f"{sheet_name}: dữ liệu trống sẵn")

    def file_creation_date(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.Cells(1, 1).Value = f"Ngày tạo file: {date_time}"

    def check_date(self, sheet_name, cell_position, num_condition):
        """
        kiểm tra ngày tháng của ô excel theo điều kiện mình cần

        Args:
            sheet_name (str): Tên sheet chứa ô cần kiểm.
            cell_position: vị trí ô cần kiểm.
            num_condition: điều kiện (cho phép trong -3,-2,-1,0,1,2,3)

        Returns:
            true/false: Đối tượng trả true hoặc false.
        """
        # Truy cập vào sheet
        sheet = self.get_sheet(sheet_name)

        # Lấy giá trị của ô cần kiểm tra
        cell_value = sheet.Range(cell_position).Value

        # Kiểm tra nếu giá trị trong ô là một ngày hợp lệ
        if isinstance(cell_value, datetime):
            # Lấy ngày hiện tại
            current_date = datetime.now().date()

            # Tính toán ngày cần so sánh dựa trên num_condition
            if num_condition == 0:
                target_date = current_date
            elif num_condition == -1:
                target_date = current_date - timedelta(days=1)
            elif num_condition == 1:
                target_date = current_date + timedelta(days=1)
            elif num_condition == -2:
                target_date = current_date - timedelta(days=2)
            elif num_condition == -3:
                target_date = current_date - timedelta(days=3)
            elif num_condition == 2:
                target_date = current_date + timedelta(days=2)
            elif num_condition == 3:
                target_date = current_date + timedelta(days=3)
            else:
                return False

            # So sánh ngày trong ô với ngày tính toán
            if cell_value.date() == target_date:
                return True
            else:
                return False
        else:
            return False


### PANDAS ###
def excel_column_to_number(column_name):
    """
    Chuyển đổi tên cột Excel (A, B, ..., Z, AA, AB, ...) sang số tương ứng.

    Args:
        column_name (str): Tên cột Excel (ví dụ: "A", "Z", "AA").

    Returns:
        int: Số thứ tự cột tương ứng.
    """
    column_name = column_name.upper()  # Đảm bảo chữ in hoa
    column_number = 0
    for char in column_name:
        # Mỗi ký tự đóng góp giá trị dựa trên vị trí của nó
        column_number = column_number * 26 + (ord(char) - ord("A") + 1)
    return int(column_number)


def save_to_excel(filtered_data, output_file):
    # Kiểm tra nếu có dữ liệu để lưu
    if filtered_data is not None:
        # Tạo tên hàng đầu với ngày giờ
        date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Tạo header với ngày giờ tạo file
        header = pd.DataFrame([[f"Ngày tạo file: {date_time}"]])

        # Mở file Excel đã định dạng sẵn
        wb = load_workbook(output_file)
        ws = (
            wb.active
        )  # Giả sử sheet bạn cần là sheet đầu tiên, bạn có thể thay đổi nếu cần

        # Thêm dòng header vào ô A1 (hàng đầu tiên)
        ws["A1"] = f"Ngày tạo file: {date_time}"

        # Đặt dữ liệu bắt đầu từ hàng thứ 3 (hàng 2 là header)
        for r_idx, row in enumerate(
            filtered_data.values, start=3
        ):  # start=3 để bắt đầu từ hàng thứ 3
            for c_idx, value in enumerate(row, start=1):  # start=1 để bắt đầu từ cột A
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Lưu lại file mới
        wb.save(output_file)
        print(f"Dữ liệu đã được lưu vào {output_file}")
    else:
        print("Không có dữ liệu để lưu.")


def excel_to_image(excel_file, output_image_file):
    # Đọc dữ liệu từ Excel
    df = pd.read_excel(excel_file)

    # Thay thế NaN hoặc None bằng chuỗi trống hoặc giá trị mặc định
    df = df.fillna("")

    # Thay thế các cột có tên "Unnamed" bằng chuỗi trống
    df.columns = df.columns.str.replace("Unnamed.*", "", regex=True)

    # Thay thế các giá trị "Unnamed" trong các ô dữ liệu bằng chuỗi trống
    df = df.apply(lambda col: col.apply(lambda x: "" if "Unnamed" in str(x) else x))

    # Tạo một plot từ dữ liệu
    fig, ax = plt.subplots(
        figsize=(16, 10)
    )  # Thiết lập kích thước ảnh (có thể thay đổi)
    ax.axis("tight")
    ax.axis("off")  # Tắt trục

    # Tạo bảng trong matplotlib
    table = ax.table(
        cellText=df.values, colLabels=df.columns, cellLoc="center", loc="center"
    )

    # Định dạng cột đầu tiên (chữ đậm, màu nền)
    for (i, j), cell in table.get_celld().items():
        if i == 1:  # Nếu là hàng tiêu đề
            cell.set_fontsize(40)
            cell.set_text_props(
                weight="bold", color="black", fontfamily="Times New Roman"
            )  # Đổi màu chữ thành trắng và chữ đậm
            cell.set_facecolor("#4CAF50")  # Màu nền (màu xanh lá cây)
        else:
            cell.set_fontsize(26)

    # Điều chỉnh độ rộng của bảng (thu nhỏ độ rộng của cột)
    for key, cell in table.get_celld().items():
        cell.set_width(
            0.3
        )  # Điều chỉnh độ rộng của từng ô trong bảng (0.1 là kích thước nhỏ hơn)

    # Điều chỉnh layout để loại bỏ phần trắng xung quanh
    plt.subplots_adjust(top=0.99, bottom=0.01)  # Giảm không gian xung quanh bảng

    # Điều chỉnh độ cao hàng
    for i, key in enumerate(table.get_celld().keys()):
        cell = table.get_celld()[key]
        # if key[0] != 0:  # Không thay đổi hàng tiêu đề
        cell.set_height(0.10)  # Điều chỉnh độ cao hàng (giá trị có thể thay đổi)

    max_col_width = 1.0
    # Điều chỉnh chiều rộng tối đa của cột và tránh cắt bớt nội dung
    for i, col in enumerate(df.columns):
        max_width = max(
            [len(str(val)) for val in df[col].values] + [len(col)]
        )  # Tính chiều rộng tối đa của cột
        max_width = min(
            max_width * 0.4, max_col_width
        )  # Giới hạn chiều rộng tối đa của cột
        table.auto_set_column_width(col=[i])  # Cố gắng thu hẹp cột
        table._cells[(0, i)].set_width(max_width)  # Đặt chiều rộng cột

    # Lưu bảng dưới dạng ảnh
    plt.savefig(output_image_file, format="png", bbox_inches="tight")
    plt.close()


def clear_data(sheet, start_row, start_col, end_col):
    """
    Xóa dữ liệu trong sheet, từ cột bắt đầu đến cột kết thúc, từ dòng bắt đầu đến dòng cuối cùng.

    Args:
        sheet (Worksheet): Sheet cần xóa dữ liệu.
        start_row (int): Dòng bắt đầu xóa.
        start_col (str): Cột bắt đầu xóa (ký tự, ví dụ: 'A').
        end_col (str): Cột kết thúc xóa (ký tự, ví dụ: 'Z').
    """
    # Xác định cột bắt đầu và cột kết thúc theo chỉ số
    start_col_idx = sheet[start_col + str(start_row)].column
    end_col_idx = sheet[end_col + str(start_row)].column

    # Tìm dòng cuối cùng trong sheet
    max_row = sheet.max_row

    # Lặp qua từng dòng và từng cột trong phạm vi cần xóa
    for row in range(start_row, max_row + 1):
        for col in range(start_col_idx, end_col_idx + 1):
            sheet.cell(row=row, column=col).value = None


def filter_ft_wo_ton(file_path, sheet_name, group_name):
    # Đọc dữ liệu từ Excel, chỉ lấy các cột từ B đến O và bắt đầu từ hàng 4
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=3, usecols="B:O")

    # df.columns = df.columns.str.strip()

    df.iloc[:, 3] = df.iloc[:, 3].astype(int)
    df.iloc[:, 8] = df.iloc[:, 8].astype(int)

    # Thêm điều kiện lọc vào DataFrame
    filtered_df = df[
        (df.iloc[:, 0] == group_name)  # Cột Mã cụm
        & (
            (df.iloc[:, 3] != 0) | (df.iloc[:, 8] != 0)
        )  # Cột Tổng == "WO tồn" hoặc "WO tồn quá hạn"
    ]

    if not filtered_df.empty:
        # lấy dữ liệu theo tên cột
        filtered_data = filtered_df.iloc[:, 1]  # cột User thực hiện
        filtered_data = filtered_data.fillna("")
        # filtered_data = filtered_data.applymap(str)
        return filtered_data
    else:
        print(f"'{group_name}' không có wo tồn")
        return None


def copy_history_KV3(excel_tool_manager: ExcelManager):
    """
    hàm này để copy dữ liệu cột N sang cột N-1 trong sheet KV3
    """
    excel_tool_manager
    try:
        excel_tool_manager.open_file()
        sheet = excel_tool_manager.get_sheet("KV3")
        sheet.Range("C2").Formula = get_DATE_fomulas()
        try:
            # copy cột N-3 sáng cột N-4
            temp = excel_tool_manager.copy_result_data(
                "KV3", start_row=5, start_col="Q", end_col="Q"
            )
            excel_tool_manager.paste_result_data(
                temp, "KV3", dest_start_row=5, dest_start_col="P"
            )

            # copy cột N-2 sáng cột N-3
            temp = excel_tool_manager.copy_result_data(
                "KV3", start_row=5, start_col="R", end_col="R"
            )
            excel_tool_manager.paste_result_data(
                temp, "KV3", dest_start_row=5, dest_start_col="Q"
            )

            # copy cột N-1 sáng cột N-2
            temp = excel_tool_manager.copy_result_data(
                "KV3", start_row=5, start_col="S", end_col="S"
            )
            excel_tool_manager.paste_result_data(
                temp, "KV3", dest_start_row=5, dest_start_col="R"
            )

            # lưu file
            excel_tool_manager.save_file()
            excel_tool_manager.close_all_file()

        except Exception as e:
            print(e)
            excel_tool_manager.close_all_file()
    except Exception as e:
        excel_tool_manager.close_all_file()
        print(e)


def get_DATE_fomulas():
    date_obj = datetime.now()
    day = date_obj.day  # Lấy ngày
    month = date_obj.month
    year = date_obj.year
    formula = f"=DATE({year},{month},{day})"
    return formula
